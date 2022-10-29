"""
Takes the xl sheets with the information to text.

Should read from input.xlsx in IO and output data to the texts files in the specified format. 

Use openpyxl to read data from the spreadsheet,

Keep all of the program inside of functions so it can be accessed as an import from "main.py"
"""

from openpyxl import load_workbook


def xlsx_to_txt() -> None:

    """
    Reads the input spreadsheet and converts each page into text.
    Saves in individual text files.
    """
    
    files = ['students.txt', 'teachers.txt', 'subjects.txt', 'rooms.txt'] #txt file name to output to

    for i in range(4): #iterates through excel sheets
        file_to_delete = open(files[i], 'w')
        file_to_delete.close() #clears txt file
        wb = load_workbook("IO\input.xlsx") #opens input.xlsx
        wb.active=wb.worksheets[i] #assigns sheet i
        sheet = wb.active #selects sheet i
        maxRow = sheet.max_row #assigns final row
        maxColumn=sheet.max_column #asssigns final column
        for x in range(2, maxRow + 1): #iterates through rows
            for y in range(1, maxColumn +1): #iterates through columns
                cell_obj = sheet.cell(row=x, column=y) #selects cell
                if cell_obj.value == '*':
                    cell_obj.value = 'w1MONp1,w1MONp2,w1MONp3,w1MONp4,w1MONp5,w1TUEp1,w1TUEp2,w1TUEp3,w1TUEp4,w1TUEp5,w1WEDp1,w1WEDp2,w1WEDp3,w1WEDp4,w1WEDp5,w1THUp1,w1THUp2,w1THUp3,w1THUp4,w1THUp5,w1FRIp1,w1FRIp2,w1FRIp3,w1FRIp4,w1FRIp5,w2MONp1,w2MONp2,w2MONp3,w2MONp4,w2MONp5,w2TUEp1,w2TUEp2,w2TUEp3,w2TUEp4,w2TUEp5,w2WEDp1,w2WEDp2,w2WEDp3,w2WEDp4,w2WEDp5,w2THUp1,w2THUp2,w2THUp3,w2THUp4,w2THUp5,w2FRIp1,w2FRIp2,w2FRIp3,w2FRIp4,w2FRIp5'
                print(cell_obj.value, end=' | ', file=open(files[i], 'a')) #prints cell
            if x < maxRow: 
                print('', file=open(files[i], 'a')) #newline

# write code in here if you only want it to be accessed when this file is run directly, not as an import
def main () -> None:
    xlsx_to_txt()

if __name__ == '__main__':
    main ()

